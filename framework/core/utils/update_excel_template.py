import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, Protection
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule, CellIsRule
from openpyxl.utils import get_column_letter

def format_excel_template(file_path=r"c:\Users\datbt20\Documents\projects\gui-testing-tool\test-data\Master_Test_Suite.xlsx"):
    wb = openpyxl.load_workbook(file_path)

    # Define Styles
    header_fill = PatternFill(start_color="1A3A4A", end_color="1A3A4A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    readonly_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid") # light grey
    wrap_align = Alignment(wrap_text=True, vertical="top")
    center_align = Alignment(horizontal="center", vertical="top")
    left_align = Alignment(horizontal="left", vertical="top")
    monospace_font = Font(name="Consolas", bold=True)
    bold_font = Font(bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    red_font = Font(color="9C0006")
    green_font = Font(color="006100")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    def format_header(ws, freeze=True):
        if freeze:
            ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align

    # 1. ACTION_LIST
    if "ACTION_LIST" in wb.sheetnames:
        ws_action = wb["ACTION_LIST"]
        format_header(ws_action, freeze=False)
        ws_action.freeze_panes = None
        ws_action.sheet_view.pane = None
        ws_action.sheet_view.selection = []
        for row in range(2, ws_action.max_row + 1):
            ws_action.cell(row=row, column=1).font = monospace_font
            ws_action.cell(row=row, column=2).alignment = wrap_align
        ws_action.column_dimensions['B'].width = 40
        # Do not protect ACTION_LIST, ensure cells are not locked
        for row in ws_action.iter_rows():
            for cell in row:
                cell.protection = Protection(locked=False)
        ws_action.protection.sheet = False

    # 2. PAGE
    if "PAGE" in wb.sheetnames:
        ws_page = wb["PAGE"]
        format_header(ws_page)
        ws_page.column_dimensions['B'].width = 60
        ws_page.conditional_formatting.add("A2:A100", FormulaRule(formula=['COUNTIF($A$2:$A$100,A2)>1'], fill=red_fill, font=red_font))

    # 3. ELEMENT
    if "ELEMENT" in wb.sheetnames:
        ws_el = wb["ELEMENT"]
        format_header(ws_el)
        ws_el.column_dimensions['C'].width = 50
        for row in range(2, ws_el.max_row + 1):
            ws_el.cell(row=row, column=3).alignment = wrap_align
        ws_el.conditional_formatting.add("A2:A1000", FormulaRule(formula=['COUNTIF($A$2:$A$1000,A2)>1'], fill=red_fill, font=red_font))
        dv_locator = DataValidation(
            type="list", 
            formula1='"css,xpath,id,name,page"', 
            allow_blank=True,
            showErrorMessage=True,
            errorStyle="stop",
            errorTitle="Dữ liệu không hợp lệ",
            error="Vui lòng chọn đúng locator type từ danh sách."
        )
        ws_el.data_validations.dataValidation = []
        ws_el.add_data_validation(dv_locator)
        dv_locator.add("B2:B1000")

    # 4. DATA_*
    for sheetname in wb.sheetnames:
        if sheetname.upper().startswith("DATA_"):
            ws_data = wb[sheetname]
            format_header(ws_data)
            type_col = None
            for col in range(1, ws_data.max_column + 1):
                if ws_data.cell(row=1, column=col).value == "test_case_type":
                    type_col = col
                    break
            if type_col:
                col_letter = get_column_letter(type_col)
                dv_type = DataValidation(
                    type="list", 
                    formula1='"pos,neg"', 
                    allow_blank=True,
                    showErrorMessage=True,
                    errorStyle="stop",
                    errorTitle="Dữ liệu không hợp lệ",
                    error="Vui lòng chọn 'pos' hoặc 'neg'."
                )
                ws_data.data_validations.dataValidation = []
                ws_data.add_data_validation(dv_type)
                dv_type.add(f"{col_letter}2:{col_letter}100")
                for row in range(2, ws_data.max_row + 1):
                    ws_data.cell(row=row, column=type_col).alignment = center_align

    # 5. TEST_CASE sheets
    for sheetname in wb.sheetnames:
        if sheetname.upper().startswith("TEST_CASE"):
            ws_tc = wb[sheetname]
            ws_tc.title = sheetname.upper()
            format_header(ws_tc)

            # Columns formatting
            ws_tc.column_dimensions['B'].width = 40 # summary
            ws_tc.column_dimensions['D'].width = 10 # step
            ws_tc.column_dimensions['E'].width = 20 # action
            ws_tc.column_dimensions['F'].width = 25 # target
            ws_tc.column_dimensions['G'].width = 30 # value
            ws_tc.column_dimensions['H'].width = 20 # expected
            ws_tc.column_dimensions['L'].width = 30 # screenshot
            
            # Apply borders and unlock cells
            for row in ws_tc.iter_rows(min_row=1, max_row=100, min_col=1, max_col=12):
                for cell in row:
                    if cell.row <= 50:
                        cell.border = thin_border
                    
                    # Protection logic
                    is_header = (cell.row == 1)
                    is_output_column = (cell.column >= 9 and cell.column <= 12)
                    if is_header or is_output_column:
                        cell.protection = Protection(locked=True)
                        if is_output_column and cell.row > 1 and cell.row <= 50:
                            cell.fill = readonly_fill
                    else:
                        cell.protection = Protection(locked=False)

            for row in range(2, 51):
                ws_tc.cell(row=row, column=1).font = bold_font
                ws_tc.cell(row=row, column=2).alignment = wrap_align
                ws_tc.cell(row=row, column=4).alignment = center_align
                ws_tc.cell(row=row, column=7).alignment = left_align

            # Dropdown type (Col C)
            dv_tc_type = DataValidation(
                type="list", 
                formula1='"pos,neg"', 
                allow_blank=True,
                showErrorMessage=True,
                errorStyle="stop",
                errorTitle="Dữ liệu không hợp lệ",
                error="Vui lòng chọn 'pos' hoặc 'neg'."
            )
            dv_tc_type.add("C2:C100")

            # Dropdown action (Col E)
            dv_action = DataValidation(
                type="list", 
                formula1="OFFSET(ACTION_LIST!$A$2,0,0,COUNTA(ACTION_LIST!$A:$A)-1,1)", 
                allow_blank=True,
                showErrorMessage=True,
                errorStyle="stop",
                errorTitle="Dữ liệu không hợp lệ",
                error="Vui lòng chọn đúng action từ danh sách ACTION_LIST."
            )
            dv_action.add("E2:E100")

            # Dropdown target (Col F)
            dv_target = DataValidation(
                type="list",
                formula1="OFFSET(ELEMENT!$A$2,0,0,COUNTA(ELEMENT!$A:$A)-1,1)",
                allow_blank=True,
                showErrorMessage=True,
                errorStyle="stop",
                errorTitle="Dữ liệu không hợp lệ",
                error="Vui lòng chọn đúng target từ danh sách element_id của sheet ELEMENT."
            )
            dv_target.add("F2:F100")
            
            ws_tc.data_validations.dataValidation = []
            ws_tc.add_data_validation(dv_tc_type)
            ws_tc.add_data_validation(dv_action)
            ws_tc.add_data_validation(dv_target)

            # Conditional formatting for Result (Col I,J,K,L)
            res_col = "I" # Default
            for col in range(9, 13):
                if ws_tc.cell(row=1, column=col).value == "[o]_test_result":
                    res_col = get_column_letter(col)
                    break

            ws_tc.conditional_formatting.add(f"{res_col}2:{res_col}100", CellIsRule(operator='equal', formula=['"Passed"'], fill=green_fill, font=green_font))
            ws_tc.conditional_formatting.add(f"{res_col}2:{res_col}100", CellIsRule(operator='equal', formula=['"Failed"'], fill=red_fill, font=red_font))

            # Enable sheet protection
            ws_tc.protection.sheet = True
            ws_tc.protection.password = '123'
            ws_tc.protection.formatCells = False
            ws_tc.protection.insertRows = False
            ws_tc.protection.deleteRows = False

    wb.save(file_path)
    print(f"Successfully formatted template: {file_path}")

if __name__ == "__main__":
    format_excel_template(r"c:\Users\datbt20\Documents\projects\gui-testing-tool\test-data\Master_Test_Suite.xlsx")
    format_excel_template(r"c:\Users\datbt20\Documents\projects\gui-testing-tool\test-data\Template_Master_Test_Suite.xlsx")
