import openpyxl

def read_test_data(file_path):
    """
    Read test data from an Excel file containing sheets:
    TEST_CASE, PAGE, ELEMENT, DATA
    Returns a dictionary with all the parsed data.
    """
    data = {
        "elements": {},
        "pages": {},
        "scenarios": [],
        "test_cases": [],
        "test_data": {}
    }
    
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        
        # 1. Read ELEMENT
        if "ELEMENT" in workbook.sheetnames:
            sheet = workbook["ELEMENT"]
            headers = [cell.value for cell in sheet[1]]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row[0]: continue
                row_data = dict(zip(headers, row))
                element_key = row_data.get("ElementKey") or row_data.get("element_id")
                locator_type = row_data.get("LocatorType") or row_data.get("locator_type")
                locator_value = row_data.get("LocatorValue") or row_data.get("locator_value")
                if element_key:
                    data["elements"][element_key] = {
                        "locator": locator_value,
                        "locator_type": locator_type
                    }

        # 2. Read PAGE
        if "PAGE" in workbook.sheetnames:
            sheet = workbook["PAGE"]
            headers = [cell.value for cell in sheet[1]]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row[0]: continue
                row_data = dict(zip(headers, row))
                page_key = row_data.get("PageKey") or row_data.get("page")
                url = row_data.get("URL") or row_data.get("url")
                page_name = row_data.get("Page Name") or page_key
                if page_key:
                    data["pages"][page_key] = {"url": url, "name": page_name}

        # 3. Read DATA (support DATA and DATA_ prefix sheets)
        # We need to map datasets index-by-index per test_case_type
        # dataset indices tracks how many rows we have added for a given test_case_type
        dataset_indices = {}
        for sheet_name in workbook.sheetnames:
            if sheet_name == "DATA" or sheet_name.startswith("DATA_"):
                sheet = workbook[sheet_name]
                sheet_key = sheet_name.lower()
                headers = [cell.value for cell in sheet[1]]
                
                # Reset counts for each sheet so we match row 1 of DATA_A with row 1 of DATA_B
                sheet_dataset_indices = {}
                
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if not row[0]: continue # DataSet is empty
                    row_data = dict(zip(headers, row))
                    dataset = row_data.get("DataType") or row_data.get("DataSet") or row_data.get("test_case_type")
                    if dataset:
                        if dataset not in data["test_data"]:
                            data["test_data"][dataset] = []
                            
                        if dataset not in sheet_dataset_indices:
                            sheet_dataset_indices[dataset] = 0
                            
                        idx = sheet_dataset_indices[dataset]
                        
                        if idx >= len(data["test_data"][dataset]):
                            data["test_data"][dataset].append({"_flat": {}})
                            
                        # Store in sheet-specific namespace
                        data["test_data"][dataset][idx][sheet_key] = row_data
                        # Store flat for backward compatibility
                        data["test_data"][dataset][idx]["_flat"].update(row_data)
                        
                        sheet_dataset_indices[dataset] += 1

        # 4. Read TEST_CASE
        grouped_scenarios = {}
        current_tc_id = None
        current_dataset = None
        
        if "TEST_CASE" in workbook.sheetnames:
            sheet = workbook["TEST_CASE"]
            headers = [cell.value for cell in sheet[1]]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # if row is completely empty, skip
                if not any(row):
                    continue
                    
                row_data = dict(zip(headers, row))
                
                # Check if it's a new TC
                tc_id = row_data.get("TC_ID") or row_data.get("tc-id")
                if tc_id and str(tc_id).strip():
                    current_tc_id = str(tc_id).strip()
                    dataset_val = row_data.get("Data") or row_data.get("DataType") or row_data.get("DataSet") or row_data.get("type")
                    if dataset_val and str(dataset_val).strip():
                        current_dataset = str(dataset_val).strip()
                    else:
                        current_dataset = None
                    
                    if current_tc_id not in grouped_scenarios:
                        grouped_scenarios[current_tc_id] = {
                            "tc_id": current_tc_id,
                            "summary": row_data.get("Summary") or row_data.get("summary") or "",
                            "run": True,
                            "dataset": current_dataset,
                            "steps": []
                        }
                        
                if current_tc_id:
                    # check if this row has a step
                    step_no = row_data.get("Step") or row_data.get("step")
                    if step_no and str(step_no).strip():
                        # Add step to the current TC
                        # Notice we map headers from the image to standard internal keys
                        step_data = {
                            "TestCaseID": current_tc_id,
                            "StepNo": step_no,
                            "Action": row_data.get("Action") or row_data.get("action"),
                            "TargetElement": row_data.get("Target") or row_data.get("target"),
                            "DataColumn": row_data.get("Data") or row_data.get("value"),
                            "Expected": row_data.get("Expected") or row_data.get("expected")
                        }
                        grouped_scenarios[current_tc_id]["steps"].append(step_data)
                        
        data["test_cases"] = list(grouped_scenarios.values())

        return data
    except Exception as e:
        print(f"[Error] Failed to read Excel file '{file_path}': {e}")
        return data
