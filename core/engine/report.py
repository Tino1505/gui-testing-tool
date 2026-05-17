import os
import time
from jinja2 import Template

TEMPLATE_STR = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>GUI Testing Report</title>
    <style>
        body { font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background-color: #f4f7f6; color: #333; margin: 0; padding: 20px; }
        .container { max-width: 1200px; margin: auto; background: white; padding: 20px; border-radius: 8px; box-shadow: 0 4px 6px rgba(0,0,0,0.1); }
        h1 { text-align: center; color: #2c3e50; }
        .summary { display: flex; justify-content: space-around; background: #ecf0f1; padding: 15px; border-radius: 8px; margin-bottom: 20px; }
        .summary div { text-align: center; font-size: 1.2em; font-weight: bold; }
        .pass-text { color: #27ae60; }
        .fail-text { color: #c0392b; }
        .tc-card { border: 1px solid #ddd; margin-bottom: 20px; border-radius: 8px; overflow: hidden; }
        .tc-header { background-color: #34495e; color: white; padding: 10px 15px; display: flex; justify-content: space-between; align-items: center; }
        .tc-header h3 { margin: 0; }
        .status-pass { background-color: #27ae60; color: white; padding: 4px 8px; border-radius: 4px; font-weight: bold; font-size: 0.9em; }
        .status-fail { background-color: #c0392b; color: white; padding: 4px 8px; border-radius: 4px; font-weight: bold; font-size: 0.9em; }
        .tc-body { padding: 15px; display: flex; gap: 20px; }
        .steps-container { flex: 1; }
        table { width: 100%; border-collapse: collapse; }
        th, td { padding: 8px 12px; border: 1px solid #ddd; text-align: left; font-size: 0.95em; }
        th { background-color: #f9f9f9; color: #555; }
        .screenshot-container { width: 250px; display: flex; flex-direction: column; align-items: center; justify-content: center; border-left: 1px solid #ddd; padding-left: 20px; }
        .screenshot { max-width: 100%; max-height: 200px; cursor: pointer; border: 1px solid #ccc; border-radius: 4px; }
        .screenshot:hover { opacity: 0.8; }
        .modal { display: none; position: fixed; z-index: 1; left: 0; top: 0; width: 100%; height: 100%; background-color: rgba(0,0,0,0.8); }
        .modal-content { margin: auto; display: block; max-width: 90%; max-height: 90%; margin-top: 2%; }
        .close { position: absolute; top: 15px; right: 35px; color: #f1f1f1; font-size: 40px; font-weight: bold; cursor: pointer; }
    </style>
</head>
<body>
    <div class="container">
        <h1>GUI Testing Report</h1>
        
        <div class="summary">
            <div>Total TC: {{ total }}</div>
            <div class="pass-text">Passed: {{ passed }}</div>
            <div class="fail-text">Failed: {{ failed }}</div>
            <div>Time: {{ time }}</div>
        </div>
        
        <div style="text-align: center; margin-bottom: 20px; display: flex; justify-content: center; gap: 15px; flex-wrap: wrap;">
            <a href="Master_Test_Suite_Backup.xlsx" download style="background-color: #27ae60; color: white; padding: 10px 20px; text-decoration: none; border-radius: 5px; font-weight: bold; display: inline-block; box-shadow: 0 2px 4px rgba(0,0,0,0.2);">
                📊 Download Excel Result
            </a>
            <a href="#" onclick="downloadHTML()" style="background-color: #3498db; color: white; padding: 10px 20px; text-decoration: none; border-radius: 5px; font-weight: bold; display: inline-block; box-shadow: 0 2px 4px rgba(0,0,0,0.2);">
                🌐 Download HTML Report
            </a>
            <a href="#" onclick="window.print()" style="background-color: #9b59b6; color: white; padding: 10px 20px; text-decoration: none; border-radius: 5px; font-weight: bold; display: inline-block; box-shadow: 0 2px 4px rgba(0,0,0,0.2);">
                🖨️ Export as PDF
            </a>
        </div>

        {% for tc in results %}
        <div class="tc-card">
            <div class="tc-header">
                <h3>{{ tc.tc_id }}: {{ tc.summary }}</h3>
                {% if tc.passed %}
                    <span class="status-pass">PASS</span>
                {% else %}
                    <span class="status-fail">FAIL</span>
                {% endif %}
            </div>
            <div class="tc-body">
                <div class="steps-container">
                    <table>
                        <thead>
                            <tr>
                                <th>Step</th>
                                <th>Action</th>
                                <th>Target</th>
                                <th>Data</th>
                                <th>Status</th>
                                <th>Message</th>
                            </tr>
                        </thead>
                        <tbody>
                            {% for step in tc.steps %}
                            <tr>
                                <td>{{ step.step }}</td>
                                <td><b>{{ step.action }}</b></td>
                                <td><code>{{ step.target }}</code></td>
                                <td>{{ step.data }}</td>
                                <td>
                                    {% if step.passed %}
                                        <span style="color: #27ae60; font-weight: bold;">PASS</span>
                                    {% else %}
                                        <span style="color: #c0392b; font-weight: bold;">FAIL</span>
                                    {% endif %}
                                </td>
                                <td>{{ step.message }}</td>
                            </tr>
                            {% endfor %}
                        </tbody>
                    </table>
                </div>
                <div class="screenshot-container">
                    {% if tc.screenshot %}
                        <p style="margin-top: 0; font-size: 0.9em; color: #777;">Evidence</p>
                        <img class="screenshot" src="{{ tc.screenshot }}" alt="Screenshot" onclick="openModal(this.src)">
                    {% else %}
                        <p>No Screenshot</p>
                    {% endif %}
                </div>
            </div>
        </div>
        {% endfor %}
    </div>

    <!-- The Modal -->
    <div id="myModal" class="modal">
      <span class="close" onclick="closeModal()">&times;</span>
      <img class="modal-content" id="img01">
    </div>

    <script>
        function openModal(src) {
            document.getElementById("myModal").style.display = "block";
            document.getElementById("img01").src = src;
        }
        function closeModal() {
            document.getElementById("myModal").style.display = "none";
        }
        function downloadHTML() {
            var content = document.documentElement.outerHTML;
            var blob = new Blob([content], {type: 'text/html'});
            var url = window.URL.createObjectURL(blob);
            var a = document.createElement('a');
            a.href = url;
            a.download = 'Standalone_GUI_Test_Report_' + new Date().getTime() + '.html';
            a.click();
            window.URL.revokeObjectURL(url);
        }
    </script>
</body>
</html>
"""

def generate_report(results, report_dir="reports"):
    if not os.path.exists(report_dir):
        os.makedirs(report_dir)
        
    total = len(results)
    passed = sum(1 for r in results if r.get('passed'))
    failed = total - passed
    current_time = time.strftime("%Y-%m-%d %H:%M:%S")
    
    # Embed screenshot paths as base64 so HTML is fully standalone
    import base64
    html_results = []
    for tc in results:
        tc_copy = tc.copy()
        if tc_copy.get("screenshot"):
            filename = os.path.basename(tc_copy["screenshot"])
            try:
                with open(tc_copy["screenshot"], "rb") as img_file:
                    b64_str = base64.b64encode(img_file.read()).decode('utf-8')
                tc_copy["screenshot"] = f"data:image/png;base64,{b64_str}"
            except Exception as e:
                # Fallback to relative path if base64 conversion fails
                tc_copy["screenshot"] = f"screenshots/{filename}"
        html_results.append(tc_copy)

    template = Template(TEMPLATE_STR)
    
    html_content = template.render(
        results=html_results,
        total=total,
        passed=passed,
        failed=failed,
        time=current_time
    )
    
    from datetime import datetime, timezone, timedelta
    vn_tz = timezone(timedelta(hours=7))
    timestamp_str = datetime.now(vn_tz).strftime("%Y%m%d_%H%M%S")
    
    report_path = os.path.join(report_dir, f"Report_{timestamp_str}.html")
    with open(report_path, "w", encoding="utf-8") as f:
        f.write(html_content)
        
    return report_path

def update_input_excel_with_results(excel_path, results, backup_path=None):
    import openpyxl
    try:
        wb = openpyxl.load_workbook(excel_path)
        
        # Group results by base_id
        grouped_results = {}
        for tc in results:
            base_id = tc["tc_id"].split("_Iter")[0]
            if base_id not in grouped_results:
                grouped_results[base_id] = []
            grouped_results[base_id].append(tc)
            
        if "TEST_CASE" in wb.sheetnames:
            ws_tc = wb["TEST_CASE"]
            headers = [cell.value for cell in ws_tc[1]]
            
            # Ensure columns exist (fallback to old names if creating new ones, or create new ones if preferred)
            # Find the target columns, checking both old and new names
            def get_col_index(headers_list, possible_names):
                for i, h in enumerate(headers_list):
                    if h and str(h).strip().lower() in [n.lower() for n in possible_names]:
                        return i + 1
                return -1

            status_col = get_col_index(headers, ["Status", "[o] test_result"])
            if status_col == -1:
                ws_tc.cell(row=1, column=len(headers) + 1, value="[o] test_result")
                headers.append("[o] test_result")
                status_col = len(headers)

            actual_col = get_col_index(headers, ["Actual Result", "[o] observed"])
            if actual_col == -1:
                ws_tc.cell(row=1, column=len(headers) + 1, value="[o] observed")
                headers.append("[o] observed")
                actual_col = len(headers)

            screen_col = get_col_index(headers, ["Screenshot", "[o] screenshot"])
            if screen_col == -1:
                ws_tc.cell(row=1, column=len(headers) + 1, value="[o] screenshot")
                headers.append("[o] screenshot")
                screen_col = len(headers)

            dur_col = get_col_index(headers, ["Duration (s)", "[o] duration (s)"])
            if dur_col == -1:
                ws_tc.cell(row=1, column=len(headers) + 1, value="[o] duration (s)")
                headers.append("[o] duration (s)")
                dur_col = len(headers)
            
            tc_id_col = get_col_index(headers, ["TC_ID", "tc-id"])
            if tc_id_col == -1:
                raise ValueError("Could not find 'TC_ID' or 'tc-id' column in TEST_CASE sheet.")

            step_col = get_col_index(headers, ["Step", "step"])
            iter_col = get_col_index(headers, ["Iteration", "Interation", "iteration"])
            
            current_tc_id = None
            
            for row in range(2, ws_tc.max_row + 1):
                tc_val = ws_tc.cell(row=row, column=tc_id_col).value
                if tc_val and str(tc_val).strip():
                    current_tc_id = str(tc_val).strip()
                    if current_tc_id in grouped_results and iter_col != -1:
                        ws_tc.cell(row=row, column=iter_col, value=len(grouped_results[current_tc_id]))
                    
                if not current_tc_id or current_tc_id not in grouped_results:
                    continue
                    
                step_val = ws_tc.cell(row=row, column=step_col).value if step_col > 0 else None
                if step_val and str(step_val).strip():
                    step_no = str(step_val).strip()
                    
                    status_lines = []
                    actual_lines = []
                    screen_lines = []
                    dur_lines = []
                    
                    iterations = grouped_results[current_tc_id]
                    for i, tc_iter in enumerate(iterations):
                        # Find the step in this iteration
                        iter_num = i + 1
                        step_res = None
                        is_last_step = False
                        
                        for step_idx, s in enumerate(tc_iter["steps"]):
                            if str(s["step"]) == step_no:
                                step_res = s
                                is_last_step = (step_idx == len(tc_iter["steps"]) - 1)
                                break
                                
                        if step_res and is_last_step:
                            prefix = f"[iter {iter_num}] " if len(iterations) > 1 else ""
                            
                            status_lines.append(f"{prefix}{'PASS' if step_res['passed'] else 'FAIL'}")
                            
                            # actual result message
                            msg = step_res.get('message', '')
                            actual_lines.append(f"{prefix}{msg}")
                            
                            # Screenshot logic
                            screenshot = tc_iter.get('screenshot')
                            display_screenshot = ""
                            if screenshot:
                                if "reports" in screenshot:
                                    display_screenshot = screenshot.split("reports")[-1].lstrip("\\/")
                                    display_screenshot = f"\\reports\\{display_screenshot}".replace("/", "\\")
                                else:
                                    display_screenshot = screenshot
                                    
                            if display_screenshot:
                                screen_lines.append(f"{prefix}{display_screenshot}" if len(iterations) > 1 else display_screenshot)
                            
                            # Use full TC duration on the final step
                            dur = tc_iter.get('duration', '').replace('s', '')
                            dur_lines.append(f"{prefix}{dur}s")
                            
                    # Write to cells
                    ws_tc.cell(row=row, column=status_col, value="\n".join(status_lines) if status_lines else "")
                    ws_tc.cell(row=row, column=actual_col, value="\n".join(actual_lines) if actual_lines else "")
                    
                    screen_cell = ws_tc.cell(row=row, column=screen_col, value="\n".join(screen_lines) if screen_lines else "")
                    
                    # Add hyperlink if there's a valid screenshot
                    if screen_lines:
                        has_screenshot_path = any("reports" in s for s in screen_lines if s)
                        first_screenshot_iter = next((tc for tc in iterations if tc.get('screenshot')), None)
                        if has_screenshot_path and first_screenshot_iter:
                            screen_cell.hyperlink = first_screenshot_iter['screenshot']
                            screen_cell.style = "Hyperlink"
                    else:
                        if hasattr(screen_cell, "hyperlink"):
                            screen_cell.hyperlink = None
                            
                    ws_tc.cell(row=row, column=dur_col, value="\n".join(dur_lines) if dur_lines else "")
                            
        # Try to save to original file
        try:
            wb.save(excel_path)
            print(f"Excel input file updated with results: {excel_path}")
        except PermissionError:
            print(f"Warning: File {excel_path} is open in another program. Cannot update original file.")
            
        # Always save to backup path if provided
        if backup_path:
            wb.save(backup_path)
            print(f"Backup saved with results to: {backup_path}")
            
        return True
    except Exception as e:
        print(f"Failed to update input excel file: {e}")
        return False
